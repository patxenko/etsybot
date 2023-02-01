from openpyxl import load_workbook
from pathlib import Path
from datetime import datetime
import openpyxl
import os
import fnmatch


# Class to manage excel data with openpyxl.

class Copy_excel:
    def __init__(self, directory, destination):
        try:
            self.old_data = []
            path_di = Path('results/' + directory)
            if path_di.is_dir() is False:  # si no existe el directorio lo creamos
                os.mkdir('results/' + directory)
            else:
                for file_name in os.listdir('results/' + directory + '/'):
                    if fnmatch.fnmatch(file_name, '*.xlsx'):
                        wb = load_workbook('results/' + directory + "/" + file_name)
                        for s_name in wb.sheetnames:
                            current_s = wb[s_name]
                            m_row = current_s.max_row
                            for i in range(1, m_row + 1):
                                cell_obj = current_s.cell(row=i, column=2)
                                self.old_data.append(cell_obj.value)

            self.dest = 'results/' + directory + "/" + destination
            sheet_name = datetime.now().strftime("%Y_%m_%d %H_%M")
            self.wb = openpyxl.Workbook()
            sheet_namess = self.wb.sheetnames
            ss_sheet1 = self.wb[sheet_namess[0]]
            ss_sheet1.title = sheet_name
            self.wb.save(self.dest)
            self.ws = self.wb[sheet_name]
            mylist = ['titulo del producto', 'URL', 'nº reviews past 15 days', 'nº reviews past 30 days', 'Img']
            self.ws.column_dimensions['A'].width = 30
            self.ws.column_dimensions['B'].width = 40
            self.ws.column_dimensions['E'].width = 20
            self.ws.append(mylist)
            self.wb.save(self.dest)
            self.row_dest = 2
        except Exception as e:
            print("Ha ocurrido un error: pruebe a cerrar todos los ficheros que usa el bot " +str(e))
            exit()

    # Write the value in the cell defined by row_dest+column_dest
    def write_workbook(self, column_dest, value):
        c = self.ws.cell(row=self.row_dest, column=column_dest)
        c.value = value
        self.row_dest = self.row_dest + 1

    def pasafila(self):
        self.row_dest = self.row_dest + 1

    # Save excel file
    def save_excel(self):
        self.wb.save(self.dest)

    def load_info(self):
        for worksheet in self.wb:
            print(worksheet.name)

    def close_workbook(self):
        self.wb.close()
